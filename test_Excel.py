from openpyxl import load_workbook
from tkinter import *
from tkinter import filedialog

file_name_Data = filedialog.askopenfilename(title='Select file Data.xlsx', filetypes=[('Excel files', '*.xlsx')])

list_for_find=[]
wb_data = load_workbook(file_name_Data)

sheet_wb_data = wb_data['List1']
i = 2 # starting row in file DATA.xlsx

while i <= sheet_wb_data.max_row:
	list_for_find.append(str(sheet_wb_data.cell(i, 1).value))
	i += 1
print(list_for_find)

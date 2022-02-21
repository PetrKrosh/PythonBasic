# ЭТОТ ФАЙЛ ЗАПУСКАЕМ НА UKW
# ИЗ ПАПКИ  C:\Users\pkro\PycharmProjects\Excel

import os
import shutil
from tkinter.filedialog import askdirectory
from tkinter import filedialog
from openpyxl import load_workbook


# В list_files  необходимо указать список частей имен файлов для отбора
# ручной вариант. list_files = ["E_I-116192","E_I-129420.xlsx", "K_I-135487.xlsx", "H_I-164315.xlsx", "H_I-201953.xlsx", "I_I-206893.xlsx", "J_I-224500.xlsx", "J_I-229523.xlsx", "K_I-233762.xlsx", "K_I-234058.xlsx", "K_I-237445.xlsx", "L_I-239525.xlsx", "L_I-242982.xlsx", "L_I-245420.xlsx", "L_I-245508.xlsx", "L_I-252593.xlsx", "L_I-252596.xlsx", "L_I-252709.xlsx", "L_I-252711.xlsx", "L_I-254408.xlsx", "L_I-254417.xlsx", "L_I-254438.xlsx", "M_I-255670.xlsx", "L_I-255858.xlsx", "L_I-256183.xlsx", "L_I-256666.xlsx", "M_I-256784.xlsx", "M_I-258111.xlsx", "M_I-258650.xlsx", "M_I-260753.xlsx", "M_I-260757.xlsx", "M_I-261859.xlsx", "M_I-262630.xlsx", "M_I-262654.xlsx", "M_I-264540.xlsx", "M_I-265312.xlsx", "M_I-265968.xlsx", "M_I-266813.xlsx", "M_I-267374.xlsx", "M_I-267955.xlsx", "M_I-268402.xlsx", "M_I-269210.xlsx", "M_I-269276.xlsx", "M_I-270200.xlsx", "E_I-184436.xlsx", "H_I-184470.xlsx", "H_I-197923.xlsx", "K_I-235853.xlsx", "L_I-243762.xlsx", "L_I-245520.xlsx", "L_I-252594.xlsx", "M_I-253673.xlsx", "M_I-258303.xlsx", "M_I-260153.xlsx", "M_I-260754.xlsx", "M_I-265043.xlsx", "M_I-269022.xlsx", "M_I-269411.xlsx", "M_I-270204.xlsx", "E_I-184516.xlsx", "E_I-184576.xlsx", "E_I-184666.xlsx", "E_I-184670.xlsx", "I_I-219092.xlsx", "E_I-184487.xlsx", "E_I-184667.xlsx", "H_I-197874.xlsx", "I_I-209633.xlsx", "I_I-209643.xlsx", "I_I-209650.xlsx", "K_I-230641.xlsx", "K_I-235464.xlsx", "K_I-239383.xlsx", "M_I-267089.xlsx", "M_I-267163.xlsx", "M_I-267715.xlsx", "M_I-267823.xlsx", "M_I-272721.xlsx", "M_I-272743.xlsx", "M_I-272828.xlsx", "E_I-184457.xlsx", "H_I-184692.xlsx", "H_I-187167.xlsx", "I_I-210520.xlsx", "I_I-210521.xlsx", "I_I-210522.xlsx", "I_I-210523.xlsx", "I_I-210524.xlsx", "I_I-210525.xlsx", "J_I-221727.xlsx", "J_I-222148.xlsx", "J_I-222197.xlsx", "J_I-223542.xlsx", "J_I-223543.xlsx", "J_I-223544.xlsx", "J_I-223545.xlsx", "J_I-223547.xlsx", "J_I-223548.xlsx", "J_I-223549.xlsx", "J_I-223550.xlsx", "J_I-223551.xlsx", "J_I-223552.xlsx", "J_I-223553.xlsx", "J_I-223554.xlsx", "J_I-223556.xlsx", "J_I-223557.xlsx", "J_I-223558.xlsx", "J_I-223559.xlsx", "J_I-223560.xlsx", "J_I-223561.xlsx", "K_I-230901.xlsx", "K_I-231797.xlsx", "K_I-232973.xlsx", "K_I-236298.xlsx", "L_I-240893.xlsx", "L_I-243905.xlsx", "L_I-248813.xlsx", "L_I-248865.xlsx", "L_I-252187.xlsx", "M_I-258643.xlsx", "M_I-267705.xlsx", "M_I-268748.xlsx", "M_I-268849.xlsx", "M_I-272149.xlsx", "M_I-272162.xlsx", "M_I-272824.xlsx", "M_I-273653.xlsx", "M_I-268302.xlsx", "J_I-233263.xlsx", "K_I-235259.xlsx", "I_I-210529.xlsx", "E_I-155085.xlsx", "E_I-155086.xlsx", "E_I-184443.xlsx", "E_I-184498.xlsx", "E_I-184658.xlsx", "H_I-186313.xlsx", "H_I-187148.xlsx", "E_I-187868.xlsx", "E_I-187871.xlsx", "E_I-187872.xlsx", "H_I-197984.xlsx", "H_I-202904.xlsx", "H_I-203886.xlsx", "J_I-227059.xlsx", "J_I-228509.xlsx", "J_I-230545.xlsx", "J_I-230666.xlsx", "J_I-230667.xlsx", "J_I-230668.xlsx", "J_I-230694.xlsx", "J_I-223087.xlsx", "J_I-223088.xlsx", "K_I-225039.xlsx", "K_I-231773.xlsx", "L_I-241763.xlsx", "L_I-253157.xlsx", "M_I-259954.xlsx", "M_I-265266.xlsx", "M_I-269218.xlsx", "H_I-168924.xlsx", "H_I-171139.xlsx", "J_I-215694.xlsx", "J_I-223477.xlsx", "J_I-223479.xlsx", "J_I-223739.xlsx", "L_I-246982.xlsx", "L_I-248483.xlsx", "M_I-264704.xlsx", "M_I-265857.xlsx", "M_I-268750.xlsx", "M_I-269358.xlsx", "M_I-272173.xlsx", "M_I-272719.xlsx", "K_P-004475.xlsx", "L_P-006958.xlsx", "M_P-007976.xlsx", "M_P-008024.xlsx", "M_P-008043.xlsx", "M_P-008434.xlsx", "M_P-008437.xlsx", "M_P-008438.xlsx", "M_P-008439.xlsx", "M_P-008442.xlsx", "M_P-008443.xlsx", "M_P-008444.xlsx", "M_P-008520.xlsx", "M_P-008531.xlsx", "M_P-008536.xlsx", "K_I-231562.xlsx", "K_I-233441.xlsx", "L_I-239411.xlsx", "L_I-239960.xlsx", "L_I-244318.xlsx", "L_I-253066.xlsx", "M_I-259005.xlsx", "M_I-259838.xlsx", "M_I-262256.xlsx", "M_I-262809.xlsx", "M_I-267826.xlsx", "M_I-268814.xlsx", "E_I-184429.xlsx", "E_I-188382.xlsx", "E_I-106365.xlsx", "E_I-111086.xlsx", "E_I-112304.xlsx", "E_I-116169.xlsx", "E_I-116170.xlsx", "E_I-116172.xlsx", "E_I-117060.xlsx", "E_I-133238.xlsx", "E_I-163777.xlsx", "E_I-184025.xlsx", "E_I-184095.xlsx", "H_I-186434.xlsx", "I_I-200582.xlsx", "I_I-209370.xlsx", "I_I-211308.xlsx", "I_I-211310.xlsx", "J_I-222939.xlsx", "J_I-227096.xlsx", "J_I-227098.xlsx", "J_I-228510.xlsx", "K_I-234966.xlsx", "M_I-256310.xlsx", "L_I-256815.xlsx", "M_I-264515.xlsx", "M_I-268218.xlsx", "E_I-116171.xlsx", "K_I-228574.xlsx", "K_I-231820.xlsx", "K_I-231821.xlsx", "M_I-266407.xlsx", "M_I-266408.xlsx", "M_I-266433.xlsx", "M_I-272744.xlsx", "E_I-103437.xlsx", "E_I-116194.xlsx", "E_I-118289.xlsx", "E_I-118317.xlsx", "H_I-203694.xlsx", "J_I-208858.xlsx", "J_I-208859.xlsx", "J_I-208877.xlsx", "I_I-209816.xlsx", "J_I-221713.xlsx", "J_I-222259.xlsx", "J_I-222508.xlsx", "J_I-223002.xlsx", "J_I-223069.xlsx", "J_I-223183.xlsx", "J_I-223259.xlsx", "J_I-224086.xlsx", "J_I-231156.xlsx", "L_I-249712.xlsx", "M_I-265477.xlsx", "M_I-266624.xlsx", "M_I-266733.xlsx", "M_I-267242.xlsx", "M_I-270055.xlsx", "M_I-270151.xlsx", "M_I-272203.xlsx", "M_I-272410.xlsx", "E_I-120984.xlsx", "E_I-139027.xlsx", "E_I-151202.xlsx", "J_I-221790.xlsx", "J_I-222088.xlsx", "J_I-222652.xlsx", "J_I-222795.xlsx", "K_I-234197.xlsx", "K_I-234306.xlsx", "K_I-234937.xlsx", "K_I-235120.xlsx", "K_I-235397.xlsx", "L_I-246370.xlsx", "L_I-247284.xlsx", "L_I-248017.xlsx", "L_I-248018.xlsx", "L_I-248029.xlsx", "L_I-248954.xlsx", "L_I-249073.xlsx", "L_I-249081.xlsx", "L_I-249994.xlsx", "M_I-258628.xlsx", "M_I-265104.xlsx", "M_I-266701.xlsx", "M_I-266794.xlsx", "M_I-266811.xlsx", "M_I-268451.xlsx", "M_I-269778.xlsx", "M_I-269888.xlsx", "M_I-270123.xlsx", "M_I-272610.xlsx", "H_I-200488.xlsx", "J_I-222370.xlsx", "J_I-222586.xlsx", "K_I-232196.xlsx", "K_I-232216.xlsx", "M_I-266057.xlsx", "M_I-266399.xlsx", "M_I-266656.xlsx", "M_I-269813.xlsx", "M_I-269955.xlsx", "M_I-270016.xlsx", "M_I-270042.xlsx", "M_I-270096.xlsx", "M_I-270695.xlsx", "M_I-270829.xlsx", "M_I-272189.xlsx","M_I-273099.xlsx", "J_I-233264.xlsx", "K_I-235260.xlsx", "M_I-273491.xlsx", "K_I-235263.xlsx", "E_I-151537.xlsx", "J_I-223807.xlsx", "K_I-232897.xlsx", "M_I-272632.xlsx", "E_I-116195.xlsx", "E_I-116527.xlsx", "E_I-124649.xlsx", "E_I-124710.xlsx", "E_I-139028.xlsx", "E_I-150904.xlsx", "E_I-150927.xlsx", "H_I-178052.xlsx", "E_I-184027.xlsx", "E_I-184030.xlsx", "E_I-184098.xlsx", "E_I-184107.xlsx", "E_I-184108.xlsx", "I_I-186435.xlsx", "E_I-187545.xlsx", "I_I-209146.xlsx", "I_I-219080.xlsx", "K_I-230823.xlsx", "K_I-231819.xlsx", "K_I-232801.xlsx", "K_I-234112.xlsx", "M_I-267260.xlsx", "M_I-267262.xlsx", "M_I-267269.xlsx", "L_I-249392.xlsx", "L_I-250496.xlsx", "M_I-272180.xlsx", "M_I-272209.xlsx", "M_I-272579.xlsx", "J_I-220389.xlsx", "J_I-222595.xlsx", "J_I-223732.xlsx", "K_I-231735.xlsx", "K_I-235104.xlsx", "M_I-268393.xlsx", "H_P-001553.xlsx", "J_P-003092.xlsx", "J_P-003254.xlsx", "L_P-006576.xlsx", "L_P-006643.xlsx", "M_P-007829.xlsx", "M_P-008038.xlsx", "M_P-008050.xlsx", "M_P-008133.xlsx", "M_P-008251.xlsx", "M_P-008289.xlsx", "M_P-008318.xlsx", "M_P-008590.xlsx", "M_P-008595.xlsx"]

file_name_for_Selection = filedialog.askopenfilename(title='Select file for Selection from Levenkov.xlsx', filetypes=[('Excel files', '*.xlsx')])
src_folder = askdirectory(title='Select Directory SOURCE')
src_folder = src_folder + "/"
dist_folder = askdirectory(title='Select Directory DESTINATION')
dist_folder = dist_folder + "/"

# src_folder = input('Введите путь к папке, в конце укажите символ \:')
# dist_folder = input('Введите путь к Distination, в конце укажите символ \:')

#  Создаем список частей имен для отбора из файла от Левенков
# ---Start 01---
list_files = []
wb_data_for_Selection = load_workbook(file_name_for_Selection)

sheet_wb_data_for_Selection = wb_data_for_Selection['List1']
i = 2 # starting row in file DATA.xlsx

while i <= sheet_wb_data_for_Selection.max_row:
	list_files.append(str(sheet_wb_data_for_Selection.cell(i, 1).value))
	i += 1
# ---Finish 01---

# Поиск файлов, по элементам списка list_files, в src_folder и копирование только файлов в dist_folder
# ---Start 02---
for root, dirs, files in os.walk(src_folder):
    for filename in files:
        path_file = os.path.join(root, filename)
        for item in list_files:
            if path_file.find(item) != -1: # определяем вхождение части имени файла в имя файла
                # path = src_folder + filename
                shutil.copy2(path_file, dist_folder)
# ---Finish 02---